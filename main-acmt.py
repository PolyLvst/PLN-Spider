version___ = 'PLN Spider ACMT v1.7'
from uuid import uuid4
from dotenv import load_dotenv
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.drawing.image import Image
from datetime import datetime
from openpyxl import load_workbook
import json
import os
from time import time
from time import sleep
import requests
from utils import myutils

load_dotenv(override=True,verbose=True)
find_this = os.environ
URL = find_this['WEB_URL']
USER = find_this['USER_ACMT']
PASSWORD = find_this['PW_ACMT']
EXCEL_PATH = find_this['EXCEL_PATH_acmt']
ROW_AWAL = int(find_this['ROW_START_acmt'])
ROW_AKHIR = int(find_this['ROW_END_acmt'])
COL_ID = find_this['ID_COL_acmt']
COL_PHOTO = find_this['PHOTO_COL_acmt']
# -- Berapa kali untuk mencoba mencari foto saat internet tidak stabil
BANYAK_PERCOBAAN = 3
# -- Setting Foto --
desired_width = int(find_this['img_width_acmt'])
desired_height = int(find_this['img_height_acmt'])

# Sleep timer
sleep_retry_foto = 2

now__ = datetime.now()
now_time = now__.strftime('%d-%m-%Y_%H-%M-%S')
log_file_path = './logs/PLN-Spider-ACMT'+now_time+'.log'

class ACMT:
    def __init__(self,driver):
        self.created_by = find_this['creator']
        self.MyLoggerUtils = myutils.MyLoggerUtils(log_file_path=log_file_path)
        self.driver = driver
        self.snapshots_folder = './DataSnapshots'
        self.temp_folder = "./TempImages/"
        self.temp_folder_meteran = f"{self.temp_folder}meteran"
        self.temp_folder_rumah = f"{self.temp_folder}rumah"
        self.cache_img = f"{self.snapshots_folder}/cache_img.json"
        self.cache_img_rumah = f"{self.snapshots_folder}/cache_img_rumah.json"
        self.ids_path = f"{self.snapshots_folder}/cached_ids.json"
        self.checkpoint_path = f'{self.snapshots_folder}/checkpoint.json'
        self.last_logout_time = 1
        with open(f"{self.snapshots_folder}/loglastrunpath.json","w") as f:
            json.dump({"log_path":log_file_path},f)

    def Log_write(self,text,stat='info'):
        return self.MyLoggerUtils.Log_write(text,stat=stat)
    
    def __repr__(self) -> str:
        return f'\x1b[1;96m>> Created by : {self.created_by}\n>> Github : https://github.com/PolyLvst\n\x1b[1;93m@ {version___}\x1b[0m'
    
    def input_captcha(self):
        input_captcha = self.driver.find_element('xpath',"//input[contains(@class,'gwt-TextBox x-component')]")
        input_captcha.click()
        input_captcha.clear()
        captcha = input("Input the captcha shown [type ! to reset captcha] : ")
        if captcha == "!":
            return False
        input_captcha.send_keys(captcha)
        return True

    def input_login(self,user,passw):
        user.click()
        user.send_keys(USER)
        passw.click()
        passw.send_keys(PASSWORD)

    def logout_akun(self):
        current_time = time()
        elapsed_time = current_time - self.last_logout_time
        self.Log_write(f"Time since last logout : {elapsed_time}s")
        if elapsed_time <= 30:
            self.Log_write("Logout cooldown ... [30s]","warning")
            for i in range(30,0,-1):
                sleep(1)
                if i%2 == 0:
                    print(f"\r[{i}] (-_-) zzZ ",end="")
                elif i%3 == 0:
                    print(f"\r[{i}] (^-^) Zzz ",end="")
                else:
                    print(f"\r[{i}] (-_-) zZz ",end="")
            print(f"\r[0] (UWU) Logging out ",end="")
            print("\n")
        try:
            WebDriverWait(self.driver,45).until(EC.presence_of_element_located((By.XPATH,"//button[@class='x-btn-text '][contains(.,'Logout')]")))
            tombol_logout = self.driver.find_element(By.XPATH,"//button[@class='x-btn-text '][contains(.,'Logout')]")
            tombol_logout.click()
            WebDriverWait(self.driver,45).until(EC.presence_of_element_located((By.XPATH,"//button[@class='x-btn-text '][contains(.,'Yes')]")))
            tombol_yes = self.driver.find_element(By.XPATH,"//button[@class='x-btn-text '][contains(.,'Yes')]")
            tombol_yes.click()
            self.Log_write("Logged out ... ")
        except Exception:
            self.Log_write("Logout button not found","error")
            exit(1)

    def click_sidebar(self,trying=0):
        if trying >= 3:
            self.Log_write("Something went wrong final ... [Sidebar not detected]","error")
            raise
        try:
            # Overlay loading
            self.driver.find_element(By.CLASS_NAME, "GCNLWM1NBC")
            WebDriverWait(self.driver, 40).until(EC.invisibility_of_element_located((By.CLASS_NAME, "GCNLWM1NBC")))
        except Exception:
            self.Log_write("Great no overlay .. ")
        try:
            # Folder Informasi
            WebDriverWait(self.driver, 40).until(EC.presence_of_element_located((By.XPATH, "(//img[contains(@class,' x-tree3-node-joint')])[7]")))
            informasi = self.driver.find_element(By.XPATH,"(//img[contains(@class,' x-tree3-node-joint')])[7]")
            informasi.click()
            WebDriverWait(self.driver, 40).until(EC.presence_of_element_located((By.XPATH, "//span[contains(.,'History Pelanggan Prabayar')]")))
            prabayar = self.driver.find_element(By.XPATH,"//span[contains(.,'History Pelanggan Prabayar')]")
            prabayar.click()
        except Exception:
            self.Log_write("Something went wrong [Sidebar not detected]","error")
            self.Log_write("Trying to refresh it ...","error")
            self.driver.refresh()
            self.click_sidebar(trying=trying+1)
            return
            # exit(1)

    def search_pelanggan(self,id_pelanggan):
        try:
            input_pelanggan = self.driver.find_element('xpath',"//input[contains(@name,'idpel')]")
            input_pelanggan.clear()
            input_pelanggan.send_keys(id_pelanggan)
        except Exception:
            self.Log_write("Input pelanggan not found","error")
            raise
        try:
            WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH,"(//img[@class=' x-btn-image'])[2]")))
            tombol_load = self.driver.find_element('xpath',"(//img[@class=' x-btn-image'])[2]")
            tombol_load.click()
        except Exception:
            raise

    def table_filter(self,flip=False,trying=0,id_pelanggan=None):
        if trying >= 3:
            self.Log_write(f">> Error in filter blth final ... {e}","error")
            raise
        try:
            WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//div[@role='columnheader'][contains(.,'BLTH')]")))
            filter_blth = self.driver.find_element('xpath',"//div[@role='columnheader'][contains(.,'BLTH')]")
            attribut_filter = filter_blth.get_attribute("class")
            sort_asc_flag = "sort-asc" in attribut_filter
            sort_desc_flag = "sort-desc" in attribut_filter
            # Trigger when first page load
            if sort_desc_flag == False and sort_asc_flag == False:
                filter_blth.click()
                filter_blth.click()
                self.Log_write(">> Filtered BLTH desc")
            # Triggered when this iteration set to ascending
            if sort_desc_flag == False and sort_asc_flag == True:
                filter_blth.click()
                self.Log_write(">> Filtered BLTH desc")
            if flip:
                filter_blth.click()
                self.Log_write(">> Filtered BLTH asc")
            
        except Exception as e:
            self.Log_write(f">> Error in filter blth {e}","error")
            self.click_sidebar()
            self.search_pelanggan(id_pelanggan)
            self.table_filter(trying=trying+1,id_pelanggan=id_pelanggan)
            return
        try:
            WebDriverWait(self.driver, 30).until(EC.visibility_of_element_located((By.XPATH,"//div[contains(@class,'x-grid3-cell-inner x-grid3-col-blth')]")))
            element = self.driver.find_element(By.XPATH, "//div[contains(@class,'x-grid3-cell-inner x-grid3-col-blth')]")
            actions = ActionChains(self.driver)
            actions.double_click(element).perform()
        except Exception:
            self.Log_write("--> Table not reachable","error")
            raise

    def lihat_foto_rumah(self, id_pelanggan, request_session):
        WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH,f"//img[contains(@title,'PHOTO RUMAH - {id_pelanggan}')]")))
        img_element = self.driver.find_element('xpath',f"//img[contains(@title,'PHOTO RUMAH - {id_pelanggan}')]")
        image_width = img_element.get_attribute('width')
        image_height = img_element.get_attribute('height')
        if image_width > '0' or image_height > '0':
            image_source = img_element.get_attribute("src")
            try:
                response = request_session.get(image_source,timeout=30)
            except Exception as e:
                print(e)
                print(">> Trying to sleep it off, and trying again [5s]")
                sleep(5)
                response = request_session.get(image_source,timeout=30)
            if len(response.content) == 0:
                self.Log_write(f">> Warning image size is 0 bytes ...")
                return False
            else:
                self.Log_write(f">> Response image rumah [OK] : {response.status_code} | {image_source}")
                return response
        else:
            return False

    def lihat_foto(self,id_pelanggan,request_session):
        trying = 1
        final_image_source = None
        try_flipping_filter = True
        while True:
            self.Log_write(f'--> ID : {id_pelanggan}, Percobaan ke : {trying}')
            # will exit after +1 trying
            if trying > BANYAK_PERCOBAAN:
                # if trying >= BANYAK_PERCOBAAN+1:
                self.Log_write('--> Bad connection [Refreshing]','error')
                raise
            try:
                # Frame foto iframe
                try:
                    img_frames = WebDriverWait(self.driver, 15).until(EC.visibility_of_any_elements_located((By.CLASS_NAME,"gwt-Frame")))
                except Exception:
                    self.Log_write(f">> Frames not found possible error is table not clicked ...","error")
                    self.table_filter(id_pelanggan=id_pelanggan)
                    img_frames = WebDriverWait(self.driver, 15).until(EC.visibility_of_any_elements_located((By.CLASS_NAME,"gwt-Frame")))
                        
                for fr_num,frame in enumerate(img_frames):
                    self.Log_write(f">> Switching to frame : {fr_num}")
                    self.driver.switch_to.frame(frame)
                    WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH,"//img[contains(@id,'image')]")))
                    img_element = self.driver.find_element('xpath',"//img[contains(@id,'image')]")
                    image_width = img_element.get_attribute('width')
                    image_height = img_element.get_attribute('height')
                    if image_width > '0' or image_height > '0':
                        image_source = img_element.get_attribute("src")
                        self.driver.switch_to.default_content()
                        try:
                            response = request_session.get(image_source,timeout=30)
                        except Exception as e:
                            print(e)
                            print(">> Trying to sleep it off, and trying again [5s]")
                            sleep(5)
                            response = request_session.get(image_source,timeout=30)
                        if len(response.content) == 0:
                            self.Log_write(f">> Warning image size is 0 bytes ...")
                            continue
                        self.Log_write(f">> Response image [OK] : {response.status_code} | {image_source}")
                        final_image_source = response
                        break
                    self.driver.switch_to.default_content()
                if final_image_source != None:
                    break
                if try_flipping_filter:
                    self.Log_write(f">> No image trying asc blth filter [1 time] ...")
                    self.Log_write(f">> Closing frame ...")
                    WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH,"//div[contains(@class, 'x-tool-close')]")))
                    close_button = self.driver.find_element('xpath',"//div[contains(@class, 'x-tool-close')]")
                    close_button.click()
                    self.table_filter(flip=True,id_pelanggan=id_pelanggan)
                    try_flipping_filter = False
                    continue
                self.Log_write(f">> No image")
                final_image_source = False
                break
            except Exception as e:
                trying+=1
                uid_err = uuid4()
                self.Log_write(f"Something went wrong [Trying again] photo error : {uid_err} {e}","error")
                self.driver.save_screenshot(f"./logs/Error_{uid_err}.png")
                self.driver.refresh()
                self.click_sidebar()
                self.search_pelanggan(id_pelanggan)
                self.table_filter(id_pelanggan=id_pelanggan)
        WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH,"//div[contains(@class, 'x-tool-close')]")))
        close_button = self.driver.find_element('xpath',"//div[contains(@class, 'x-tool-close')]")
        close_button.click()
        return final_image_source

    def checkpoint(self,row,no,id_pel):
        checkpoint = self.checkpoint_path
        data = {}
        # with open(checkpoint,'r') as f:
        #     data = json.load(f)
        data.update({'row_awal':ROW_AWAL,'checkpoint':{'no':no,'row_checkpoint':row,'id':id_pel}})
        with open(checkpoint,'w') as f:
            json.dump(data,f)
            self.Log_write('Updated checkpoint ..')

    def ask_checkpoint(self):
        data = {}
        nomer = 1
        row_checkpoint = ROW_AWAL
        checkpoint = self.checkpoint_path
        folder_snapshots = self.snapshots_folder
        if not os.path.exists(folder_snapshots):
            os.mkdir(folder_snapshots)
        self.clean_old_files(folder_snapshots)
        if os.path.exists(checkpoint):
            self.Log_write('--> Checkpoint found, using value from checkpoint')
            with open(checkpoint,'r') as f:
                try:
                    data:dict = json.load(f)
                except Exception:
                    self.Log_write('No value detected, fallback default')
                    return nomer,row_checkpoint
            check_idx:dict = data['checkpoint']
            nomer = check_idx.get('no')
            row_checkpoint = check_idx.get('row_checkpoint')
            return nomer,row_checkpoint
        else:
            data = {'row_awal':ROW_AWAL,'checkpoint':{'no':1,'row_checkpoint':ROW_AWAL,'id':'first time run'}}
            with open(checkpoint,'w') as f:
                json.dump(data,f)
            self.Log_write('Init checkpoint ..')
            return nomer,row_checkpoint

    def get_cached_ids(self):
        ids_path = self.ids_path
        data = {}
        if os.path.exists(ids_path):
            self.Log_write("Using cached ids ..")
            with open(ids_path,"r") as f:
                data = json.load(f)
            return data
        workbook = load_workbook(EXCEL_PATH)
        worksheet = workbook.active
        end_row = ROW_AKHIR+1
        for row in range(ROW_AWAL,end_row):
            id_pelanggan = worksheet[f'{COL_ID}{row}']
            str_pelanggan = id_pelanggan.value
            data[f"no-{str_pelanggan}"] = {"str_pelanggan":str_pelanggan,
                                        "status_value":"Init_state"}
        with open(ids_path,"w") as f:
            json.dump(data,f)
        self.Log_write("Init cached ids ..")
        workbook.close()
        return data

    def update_cache_ids(self,no_id,status_given):
        "no_id = no-1"
        ids_path = self.ids_path
        data = {}
        if os.path.exists(ids_path):
            with open(ids_path,"r") as f:
                data = json.load(f)
            data[no_id]["status_value"] = status_given
            with open(ids_path,"w") as f:
                json.dump(data,f)
            self.Log_write("Status value updated ..")
        else:
            self.Log_write("Something wrong happens [cached_ids not found]","error")
            exit(1)

    def download_photo_rumah(self, source_rumah, cur_pos, status_value):
        self.Log_write("--> Updating cache img rumah ..")
        cache_img = self.cache_img_rumah
        temp_folder_rumah = self.temp_folder_rumah
        response_rumah = source_rumah
        if source_rumah == False:
            image_path = "Tidak tersedia"
        else:
            image_name = f"image_{cur_pos}.jpg"  # You can change the file name logic as needed
            image_path = os.path.join(temp_folder_rumah, image_name)
            with open(image_path, "wb") as f:
                f.write(response_rumah.content)
        data = {}
        if os.path.exists(cache_img):
            with open(cache_img,"r") as f:
                data = json.load(f)
        with open(cache_img,"w") as f:
            data[cur_pos] = {"img":image_path,
                            "status_value":status_value}
            json.dump(data,f)
        self.Log_write("--> Cache img rumah updated ..")

    def download_photo(self,source,cur_pos,status_value):
        self.Log_write("--> Updating cache img ..")
        cache_img = self.cache_img
        temp_folder = self.temp_folder_meteran
        response = source
        if source == False:
            image_path = "Tidak tersedia"
        else:
            image_name = f"image_{cur_pos}.jpg"  # You can change the file name logic as needed
            image_path = os.path.join(temp_folder, image_name)
            with open(image_path, "wb") as f:
                f.write(response.content)
        data = {}
        if os.path.exists(cache_img):
            with open(cache_img,"r") as f:
                data = json.load(f)
        with open(cache_img,"w") as f:
            data[cur_pos] = {"img":image_path,
                            "status_value":status_value}
            json.dump(data,f)
        self.Log_write("--> Cache img updated ..")

    def save_photo(self):
        self.Log_write("Saving photos from cache ..")
        cache_img = self.cache_img
        data = {}
        if os.path.exists(cache_img):
            with open(cache_img,"r") as f:
                data = json.load(f)
        else:
            self.Log_write("Something wrong [no cache img found]","error")
            exit(1)
        workbook = load_workbook(EXCEL_PATH)
        worksheet = workbook.active
        starts_row = ROW_AWAL
        for customer_id in data:
            print(f"--> Saving {customer_id} ..")
            cur_pos = starts_row
            foto_cell=worksheet[f'{COL_PHOTO}{cur_pos}']
            image_path = data[customer_id]["img"]
            status_value = data[customer_id]["status_value"]
            if status_value == "False":
                pass
            else:
                img = Image(image_path)
                img.width = desired_width
                img.height = desired_height
                # add to worksheet and anchor next to cells
                worksheet.add_image(img, f'{COL_PHOTO}{cur_pos}')
            foto_cell.value = status_value
            starts_row += 1
        workbook.save(EXCEL_PATH)
        self.Log_write("--> Workbook updated!")
        workbook.close()

    def delete_temp(self):
        folder = self.temp_folder
        for file_img in os.listdir(folder):
            print(f"deleting -> {file_img}")
            os.remove(f"{folder}/{file_img}")
    
    def clean_old_files(self,path_to):
        max_age_seconds = 3 * 24 * 60 * 60
        old_files = []
        for file_path in os.listdir(path_to):
            file_path = os.path.join(path_to,file_path)
            file_stat = os.stat(file_path)
            current_time = time()
            # Calculate the age of the file in seconds
            file_age_seconds = current_time - file_stat.st_mtime
            # Compare the age with the maximum allowed age
            if file_age_seconds > max_age_seconds:
                # File is older than 3 days
                old_files.append(file_path)
        if not old_files:
            return
        print(f"Old files found in {path_to} ..")
        print("Use p for delete on prompt")
        cond = input(f"Delete all old files in ? {path_to} [y/n/p] ").lower()
        if cond == "n":
            print("No files deleted ..")
            return
        if cond != "y" and cond != "p":
            print("Invalid choice")
            exit(0)
        for file in old_files:
            # prompt mode if cond not "y"
            if cond != "y":
                inp = input(f"Delete {file} ? [y/n] ").lower()
                if inp == "n":
                    print(f"Skipping {file} ..")
                    continue
            # if cond == "y" then go delete all file without prompt
            os.remove(file)
            print(f"{file} has been deleted as it's more than 3 days old.","warning")

class SpiderACMT:
    def __init__(self,no_driver=False):
        if no_driver:
            self.driver = None
            self.acmt_crawler = ACMT(driver=None)
        else:
            # use the name of your firefox profile
            self.driver = myutils.WebScraperUtils.start_web_dv(profile="WebScraping")
            self.acmt_crawler = ACMT(driver=self.driver)
        self.trying = 0

    def run(self):
        self.driver.get(URL)
        self.__login()
        while self.trying < BANYAK_PERCOBAAN:
            try:
                self.__main()
                self.__cleanup()
                exit()
            except Exception as e:
                uid_err = uuid4()
                self.acmt_crawler.Log_write(f"Got error : {e}")
                self.acmt_crawler.Log_write(f"\x1b[1;35m--> [Refreshing] ({self.trying}) photo {uid_err} \x1b[0m","warning")
                self.driver.save_screenshot(f"./logs/Error_{uid_err}.png")
                self.driver.refresh()
                # try:
                #     self.acmt_crawler.logout_akun()
                # except:
                #     self.driver.save_screenshot(f"./logs/Error_{uuid4()}.png")
                #     self.acmt_crawler.Log_write("Something went wrong in logging out")
                self.acmt_crawler.last_logout_time = time()
                self.trying += 1
                continue
        self.acmt_crawler.Log_write('--> Bad connection exiting','error')
        exit(1)

    def save_photo(self):
        self.acmt_crawler.save_photo()

    def delete_temp_photo(self):
        self.acmt_crawler.delete_temp()
    
    def test_run(self,stop_at=3):
        print(f">> After testing done, you need to delete images in the {EXCEL_PATH} after test run")
        print(f">> Or the file might have duplicate image on top of other.")
        self.driver.get(URL)
        self.__login()
        self.__main(stop_at_offset = stop_at)

    def delete_snapshots(self):
        folder_snapshots = self.acmt_crawler.snapshots_folder
        self.acmt_crawler.clean_old_files(folder_snapshots)

    def __login(self):
        driver = self.driver
        acmt_crawler = self.acmt_crawler
        # Overlay
        try:
            overlay = driver.find_element(By.CLASS_NAME, "blockOverlay")
            WebDriverWait(driver, 40).until(EC.invisibility_of_element_located((By.CLASS_NAME, "blockOverlay")))
        except Exception:
            # Great no overlay
            pass
        element = WebDriverWait(driver, 35).until(EC.presence_of_element_located((By.XPATH, "//input[@name='user']")))
        input_login_user = driver.find_element('xpath',"//input[@name='user']")
        input_login_password = driver.find_element('xpath',"//input[@name='password']")
        button_login = driver.find_element('xpath',"//button[contains(.,'Login')]")
        button_recaptcha = driver.find_element('xpath',"//button[@class='x-btn-text '][contains(.,'ReCaptcha')]")
        if input_login_user and input_login_password:
            while True:
                input_login_user.clear()
                input_login_password.clear()
                acmt_crawler.input_login(input_login_user,input_login_password)
                ret_reset = acmt_crawler.input_captcha()
                button_login.click()
                try:
                    WebDriverWait(driver, 3).until(EC.alert_is_present())
                    print(">> Retry Captcha")
                    alert = driver.switch_to.alert
                    acmt_crawler.Log_write(alert.text)
                    alert.accept()
                    print(">> Alert accepted")
                    if ret_reset == False:
                        button_recaptcha.click()
                except Exception:
                    break
        else:
            acmt_crawler.Log_write('Something went wrong ! [User input not found]','error')
            driver.quit()

    def __main(self,stop_at_offset: int = 0):
        # Stop at offset 0 means it will stop at ROW_AKHIR if defined, it will stop at last checkpoint + stop_at_offset
        driver = self.driver
        acmt_crawler = self.acmt_crawler
        acmt_crawler.Log_write(str(acmt_crawler))
        session = requests.Session()
        selenium_cookies = driver.get_cookies()
        requests_cookies = {cookie['name']: cookie['value'] for cookie in selenium_cookies}
        session.cookies.update(requests_cookies)
        acmt_crawler.click_sidebar()
        acmt_crawler.Log_write('Logged in')
        nomer,row_checkpoint = acmt_crawler.ask_checkpoint()
        print(f">> Starting with row : {row_checkpoint}")
        cache_ids = acmt_crawler.get_cached_ids()
        if nomer == 1:
            splice_range = 0
        else:
            splice_range = nomer-1
        acmt_crawler.Log_write(f"\x1b[1;96m>> Total : {ROW_AKHIR-(ROW_AWAL-1)}\x1b[0m")
        # Convert the dictionary items to a list and slice it
        if stop_at_offset:
            items_list = list(cache_ids.items())[splice_range:splice_range+stop_at_offset]
        else:
            items_list = list(cache_ids.items())[splice_range:]
        # Gets the latest section of ids, from checkpoint to the end
        cache_ids = dict(items_list)
        for row in cache_ids:
            foto_status = cache_ids[row]["status_value"]
            str_pelanggan = cache_ids[row]["str_pelanggan"]
            if foto_status == 'True':
                acmt_crawler.Log_write(f'No.{nomer} Foto terdeteksi [Skipping] . . . ID : {str_pelanggan}')
                nomer+=1
                continue
            acmt_crawler.Log_write(f'No.{nomer} Mencari foto . . .')
            acmt_crawler.search_pelanggan(str_pelanggan)
            foto_rumah = acmt_crawler.lihat_foto_rumah(str_pelanggan,request_session = session)
            acmt_crawler.table_filter(id_pelanggan=str_pelanggan)
            data_foto = acmt_crawler.lihat_foto(str_pelanggan,request_session = session)
            if foto_rumah == False:
                acmt_crawler.Log_write("--> Foto rumah not found [No image] [ok] ..")
                acmt_crawler.download_photo_rumah(foto_rumah,row,"False")
            else:
                acmt_crawler.Log_write("--> Foto rumah found [saving] ..")
                acmt_crawler.download_photo_rumah(foto_rumah,row,"True")

            if data_foto == False:
                acmt_crawler.Log_write("--> Cache img updated [No image] ..")
                acmt_crawler.download_photo(data_foto,row,"False")
                acmt_crawler.update_cache_ids(row,"False")
            else:
                acmt_crawler.download_photo(data_foto,row,"True")
                acmt_crawler.update_cache_ids(row,"True")
            current_time = datetime.now().time()
            # Extract hour, minute, and second components
            hour = current_time.hour
            minute = current_time.minute
            second = current_time.second
            # Log_write("--> Workbook updated!")
            acmt_crawler.Log_write(f'--> {hour}:{minute}:{second}')
            acmt_crawler.checkpoint(row,nomer,str_pelanggan)
            acmt_crawler.Log_write(f"\x1b[1;96m>> Total left : {ROW_AKHIR-(ROW_AWAL-1)-nomer}\x1b[0m")
            nomer+=1
            self.trying = 0
        # acmt_crawler.logout_akun()
        acmt_crawler.Log_write('Webdriver flush\nExiting . . .')
        driver.quit()
        continue_to_save_and_delete_temp_images = True
        check_ids_vals = acmt_crawler.get_cached_ids()
        false_ids = {}
        for key, value in check_ids_vals.items():
            # Check if "status_value" is False for each key
            if value.get("status_value") == "False":
                acmt_crawler.Log_write(f"Status value for {key} is False.","warning")
                continue_to_save_and_delete_temp_images = False
                false_ids[key] = {
                    "str_pelanggan": value.get("str_pelanggan"),
                    "status_value": "False"
                }
        if continue_to_save_and_delete_temp_images is False:
            acmt_crawler.Log_write(f"Problematic ids : ","warning")
            acmt_crawler.Log_write(false_ids,"warning")
            acmt_crawler.Log_write(f"Cannot save to document, possible fix rerun this script but enable the main.save_photo()","warning")
            acmt_crawler.Log_write('\x1b[1;92mDone saving to snapshot ...')
            acmt_crawler.Log_write(str(acmt_crawler))
            exit()
        acmt_crawler.Log_write('\x1b[1;92mAll done ...')
        acmt_crawler.Log_write(str(acmt_crawler))
        input("Do you want to save photo? [ctrl c to abort]")
        acmt_crawler.save_photo()

    def __cleanup(self):
        choices = input("Do you want to delete temp images? [y/n]")
        if choices.lower() != "y":
            return
        acmt_crawler.Log_write('\x1b[1;92mCleaning up temp images ...')
        acmt_crawler = self.acmt_crawler
        acmt_crawler.delete_temp()

if __name__ == '__main__':
    main = SpiderACMT()
    snap = main.acmt_crawler.snapshots_folder
    cached_id = "/cached_ids.json"
    main.acmt_crawler.ids_path = f"{snap}{cached_id}"
    main.run()
    # Uncomment to use individual functionality you need

    # main.test_run(stop_at=3)

    # main_test = SpiderACMT(no_driver=True)
    # main_test.save_photo()
    # main_test.delete_temp_photo()