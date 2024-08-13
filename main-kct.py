version___ = 'PLN Spider KCT v1.0'
import logging
from dotenv import load_dotenv
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.firefox.options import Options
from openpyxl.drawing.image import Image
from selenium import webdriver
from datetime import datetime
from openpyxl import load_workbook
import json
import os
from time import time
from time import sleep
import requests

load_dotenv(verbose=True)
find_this = os.environ
URL = find_this['APP_URL']
USER = find_this['USR']
PASSWORD = find_this['PW']
EXCEL_PATH = find_this['EXCEL_PATH_KCT']
ROW_AWAL = int(find_this['ROW_START'])
ROW_AKHIR = int(find_this['ROW_END'])
COL_ID = find_this['ID_COL']
COL_PHOTO = find_this['PHOTO_COL']
# -- Berapa kali untuk mencoba mencari foto saat internet tidak stabil
BANYAK_PERCOBAAN = int(find_this['BANYAK_PERCOBAAN'])
# -- Setting Foto --
desired_width = int(find_this['img_width'])
desired_height = int(find_this['img_height'])

# Sleep timer
sleep_retry_foto = 2

now__ = datetime.now()
now_time = now__.strftime('%d-%m-%Y_%H-%M-%S')
log_file_path = './logs/PLN-Spider-KCT'+now_time+'.log'
with open("./DataSnapshots/loglastrunpath.json","w") as f:
    json.dump({"log_path":log_file_path},f)

excel_file_path = EXCEL_PATH

def show_vers():
    created_by = find_this['creator']
    return f'\x1b[1;96m>> Created by : {created_by}\n>> Github : https://github.com/PolyLvst\n\x1b[1;93m@ {version___}\x1b[0m'
# ------------- Selenium web driver ------------ #
def start_web_dv(profile="default"):
    options = Options()
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0"
    options.set_preference("general.useragent.override", user_agent)
    options.set_preference("network.trr.mode", 2)
    options.set_preference("network.trr.uri", "https://mozilla.cloudflare-dns.com/dns-query")
    profile_path = os.path.join(os.getenv('APPDATA'), 'Mozilla', 'Firefox', 'Profiles')
    # List all directories in the Profiles folder
    profiles = [d for d in os.listdir(profile_path) if os.path.isdir(os.path.join(profile_path, d))]
    path_prof = None 
    for path_p in profiles:
        if profile in path_p:
            path_prof = os.path.join(profile_path, path_p)
            break
    if path_prof:
        Log_write(f"-- Using {profile} profile")
        options.add_argument("-profile")
        options.add_argument(path_prof)
    driver = webdriver.Firefox(options=options)
    return driver

def Log_write(text,stat='info'):
    """Available params ->>\ndebug,info,warning,error,critical"""
    text = str(text)
    log_filename = log_file_path
    logging.basicConfig(filename=log_filename, filemode='w', format='%(asctime)s - %(levelname)s - %(message)s', level=logging.DEBUG)
    # Set the logging level for the selenium logger to WARNING
    logging.getLogger('selenium').setLevel(logging.WARNING)
    # Set the logging level for the webdriver logger to WARNING
    logging.getLogger('webdriver').setLevel(logging.WARNING)
    # Set the logging level, to prevent unwanted message showing in log file
    logging.getLogger('urllib3.connectionpool').setLevel(logging.WARNING)
    print(text)
    text = text.replace('\n',' ')
    # Map the level string to a logging level constant
    level_map = {'debug': logging.DEBUG,
                 'info': logging.INFO,
                 'warning': logging.WARNING,
                 'error': logging.ERROR,
                 'critical': logging.CRITICAL}
    log_level = level_map.get(stat.lower(), logging.INFO)
    logging.log(log_level,text)

def input_login(user,passw,btn):
    user.click()
    user.send_keys(USER)
    passw.click()
    passw.send_keys(PASSWORD)
    input_captcha = driver.find_element('id',"x-auto-3-input")
    input_captcha.click()
    captcha = input("Input the captcha shown : ")
    input_captcha.send_keys(captcha)
    btn.click()

def logout_akun():
    current_time = time()
    last_logout_time = 5
    elapsed_time = current_time - last_logout_time
    Log_write(f"Time since last logout : {elapsed_time}s")
    if elapsed_time <= 30:
        Log_write("Logout cooldown ... [30s]","warning")
        for i in range(30,0,-1):
            sleep(1)
            if i%2 == 0:
                print(f"\r[{i}] (-_-) zzZ ",end="")
            elif i%3 == 0:
                print(f"\r[{i}] (^-^) Zzz ",end="")
            else:
                print(f"\r[{i}] (-_-) zZz ",end="")
        print("\n")
    try:
        element = WebDriverWait(driver,45).until(EC.presence_of_element_located((By.XPATH,"//div[@class='GCNLWM1ON'][contains(.,'Logout')]")))
        tombol_logout = driver.find_element(By.XPATH,"//div[@class='GCNLWM1ON'][contains(.,'Logout')]")
        tombol_logout.click()
        Log_write("Logged out ... ")
    except:
        Log_write("Logout button not found","error")
        exit(1)
    
def delete_temp():
    folder = "./TempImages"
    for file_img in os.listdir(folder):
        Log_write(f"deleting -> {file_img}")
        os.remove(f"{folder}/{file_img}")

def click_sidebar():
    try:
        # Overlay loading
        overlay = driver.find_element(By.CLASS_NAME, "GCNLWM1NBC")
        WebDriverWait(driver, 40).until(EC.invisibility_of_element_located((By.CLASS_NAME, "GCNLWM1NBC")))
    except:
        Log_write("Great no overlay .. ")
    try:
        # Folder KCT info Baca KCT
        element = WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.XPATH, "//span[@class='GCNLWM1OLB'][contains(.,'Info Baca KCT')]")))
        informasi_TUL = driver.find_element(By.XPATH,"//span[@class='GCNLWM1OLB'][contains(.,'Info Baca KCT')]")
        informasi_TUL.click()
    except:
        Log_write("Something went wrong [Sidebar not detected]")
        # exit(1)
        raise Exception

def search_pelanggan(id_pelanggan):
    try:
        input_pelanggan = driver.find_element('id',"x-auto-38-input")
        input_pelanggan.clear()
        input_pelanggan.send_keys(id_pelanggan)
    except:
        Log_write("Input pelanggan not found","error")
    try:
        sleep(sleep_retry_foto)
        # tombol_cari = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,'/html/body/div[2]/div/div[1]/div[2]/div/div[2]/div[2]/div[1]/div/div[2]/div[1]/form/div/div[2]/div[1]/div/div/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/div[1]/fieldset/div/div/div/div[2]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[1]')))
        tombol_cari = driver.find_element('xpath',"//div[@class='GCNLWM1ON'][contains(.,'Cari')]")
        tombol_cari.click()
    except:
        try:
            element = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@class='GCNLWM1ON'][contains(.,'OK')]")))
            tombol_error_google = driver.find_element('xpath',"//div[@class='GCNLWM1ON'][contains(.,'OK')]")
            tombol_error_google.click()
            Log_write("Google gwt error closed","warning")
        except:
            Log_write("Something wrong happens [search_pelanggan]","error")
        # //div[@class='GCNLWM1FQ'][contains(.,'Problem detected : com.google.gwt.user.client.rpc.StatusCodeException: 0')]
        # //div[@class='GCNLWM1ON'][contains(.,'OK')]
        # exit(1)
        raise Exception

def lihat_foto(id_pelanggan,nomer):
    try:
        actions = ActionChains(driver)
        foto_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,"//img[contains(@class,'gwt-Image')]")))
        foto = driver.find_element(By.XPATH,"//img[contains(@class,'gwt-Image')]")
        driver.execute_script("arguments[0].scrollIntoView();", foto)
        actions.move_to_element(foto).perform()
    except:
        Log_write(f'--> ID : {id_pelanggan}')
        Log_write("--> Foto tidak tersedia","warning")
        return False
    trying = 1
    final_image_source = ''
    while True:
        Log_write(f'--> ID : {id_pelanggan}, Percobaan ke : {trying}')
        # will exit after +1 trying
        if trying > BANYAK_PERCOBAAN:
            # if trying >= BANYAK_PERCOBAAN+1:
            Log_write('--> Bad connection [Logout & Exit]','error')
            raise Exception

        # Foto bawah
        img_element_1 = driver.find_element('xpath',"//img[contains(@class,'gwt-Image')]")
        image_source_1 = img_element_1.get_attribute("src")
        try:
            response = requests.get(image_source_1,timeout=3)
            if response.ok:
                final_image_source = response
                break
            else:
                final_image_source = False
                break
        except:
            trying+=1
            search_pelanggan(id_pelanggan)
    return final_image_source

def clean_old_files(path_to):
    max_age_seconds = 3 * 24 * 60 * 60
    old_files = []
    for file_path in os.listdir(path_to):
        file_path = os.path.join(path_to,file_path)
        file_stat = os.stat(file_path)
        current_time = time()
        # Calculate the age of the file in seconds
        file_age_seconds = current_time - file_stat.st_mtime
        # Compare the age with the maximum allowed age
        if file_age_seconds > max_age_seconds:
            # File is older than 3 days
            old_files.append(file_path)
    if not old_files:
        return
    Log_write("Old files found in DataSnapshots ..")
    print("Use p for delete on prompt")
    cond = input("Delete all old files in DataSnapshots? [y/n/p] ").lower()
    if cond == "n":
        Log_write("No files deleted ..")
        return
    if cond != "y" and cond != "p":
        Log_write("Invalid choice")
        exit(0)
    for file in old_files:
        # prompt mode if cond not "y"
        if cond != "y":
            inp = input(f"Delete {file} ? [y/n] ").lower()
            if inp == "n":
                Log_write(f"Skipping {file} ..")
                continue
        # if cond == "y" then go delete all file without prompt
        os.remove(file)
        Log_write(f"{file} has been deleted as it's more than 3 days old.","warning")

def checkpoint(row,no,id_pel):
    checkpoint = './DataSnapshots/checkpoint.json'
    data = {}
    # with open(checkpoint,'r') as f:
    #     data = json.load(f)
    data.update({'row_awal':ROW_AWAL,'checkpoint':{'no':no,'row_checkpoint':row,'id':id_pel}})
    with open(checkpoint,'w') as f:
        json.dump(data,f)
        Log_write('Updated checkpoint ..')

def ask_checkpoint():
    data = {}
    nomer = 1
    row_awal = ROW_AWAL
    checkpoint = './DataSnapshots/checkpoint.json'
    if not os.path.exists('./DataSnapshots'):
        os.mkdir('./DataSnapshots')
    clean_old_files('./DataSnapshots')
    if os.path.exists(checkpoint):
        Log_write('--> Checkpoint found, using value from checkpoint')
        with open(checkpoint,'r') as f:
            try:
                data:dict = json.load(f)
            except:
                Log_write('No value detected, fallback default')
                return nomer,row_awal
        check_idx:dict = data['checkpoint']
        nomer = check_idx.get('no')
        row_awal = check_idx.get('row_checkpoint')
        return nomer,row_awal
    else:
        data = {'row_awal':ROW_AWAL,'checkpoint':{'no':1,'row_checkpoint':ROW_AWAL,'id':'first time run'}}
        with open(checkpoint,'w') as f:
            json.dump(data,f)
        Log_write('Init checkpoint ..')
        return nomer,row_awal

def get_cached_ids():
    ids_path = "./DataSnapshots/cached_ids.json"
    data = {}
    if os.path.exists(ids_path):
        Log_write("Using cached ids ..")
        with open(ids_path,"r") as f:
            data = json.load(f)
        return data
    workbook = load_workbook(excel_file_path)
    worksheet = workbook.active
    end_row = ROW_AKHIR+1
    for row in range(ROW_AWAL,end_row):
        id_pelanggan = worksheet[f'{COL_ID}{row}']
        str_pelanggan = id_pelanggan.value
        data[f"no-{str_pelanggan}"] = {"str_pelanggan":str_pelanggan,
                                       "status_value":"Init_state"}
    with open(ids_path,"w") as f:
        json.dump(data,f)
    Log_write("Init cached ids ..")
    workbook.close()
    return data

def update_cache_ids(no_id,status_given):
    "no_id = no-1"
    ids_path = "./DataSnapshots/cached_ids.json"
    data = {}
    if os.path.exists(ids_path):
        with open(ids_path,"r") as f:
            data = json.load(f)
        data[no_id]["status_value"] = status_given
        with open(ids_path,"w") as f:
            json.dump(data,f)
        Log_write("Status value updated ..")
    else:
        Log_write("Something wrong happens [cached_ids not found]","error")
        exit(1)

def download_photo(source,cur_pos,status_value):
    cache_img = "./DataSnapshots/cache_img.json"
    temp_folder = f"./TempImages/"
    response = source
    if source == False:
        image_path = "Tidak tersedia"
    else:
        image_name = f"image_{cur_pos}.jpg"  # You can change the file name logic as needed
        image_path = os.path.join(temp_folder, image_name)
        with open(image_path, "wb") as f:
            f.write(response.content)
    data = {}
    if os.path.exists(cache_img):
        with open(cache_img,"r") as f:
            data = json.load(f)
    with open(cache_img,"w") as f:
        data[cur_pos] = {"img":image_path,
                        "status_value":status_value}
        json.dump(data,f)
    Log_write("--> Cache img updated ..")

def save_photo():
    Log_write("Saving photos from cache ..")
    cache_img = "./DataSnapshots/cache_img.json"
    data = {}
    if os.path.exists(cache_img):
        with open(cache_img,"r") as f:
            data = json.load(f)
    else:
        Log_write("Something wrong [no cache img found]","error")
        exit(1)
    workbook = load_workbook(EXCEL_PATH)
    worksheet = workbook.active
    starts_row = ROW_AWAL
    for customer_id in data:
        Log_write(f"--> Saving {customer_id} ..")
        cur_pos = starts_row
        foto_cell=worksheet[f'{COL_PHOTO}{cur_pos}']
        image_path = data[customer_id]["img"]
        status_value = data[customer_id]["status_value"]
        if status_value == "False":
            pass
        else:
            img = Image(image_path)
            img.width = desired_width
            img.height = desired_height
            # add to worksheet and anchor next to cells
            worksheet.add_image(img, f'{COL_PHOTO}{cur_pos}')
        foto_cell.value = status_value
        starts_row += 1
    workbook.save(EXCEL_PATH)
    Log_write("--> Workbook updated!")
    workbook.close()

def main():
    driver.get(URL)
    # Overlay
    overlay = driver.find_element(By.CLASS_NAME, "blockOverlay")
    WebDriverWait(driver, 40).until(EC.invisibility_of_element_located((By.CLASS_NAME, "blockOverlay")))

    element = WebDriverWait(driver, 35).until(EC.presence_of_element_located((By.ID, "x-auto-1-input")))
    input_login_user = driver.find_element('id',"x-auto-1-input")
    input_login_password = driver.find_element('id',"x-auto-2-input")
    button_login = driver.find_element('xpath',"//div[@class='GCNLWM1ON'][contains(.,'Login')]")
    if input_login_user and input_login_password:
        # sleep(5)
        input_login(input_login_user,input_login_password,button_login)
    else:
        Log_write('Something went wrong ! [User input not found]','error')
        driver.quit()
        exit(1)
    click_sidebar()
    Log_write('Logged in')
    nomer,row_awal = ask_checkpoint()
    cache_ids = get_cached_ids()
    if nomer == 1:
        splice_range = 0
    else:
        splice_range = nomer-1
    Log_write(f"\x1b[1;96m>> Total : {ROW_AKHIR-ROW_AWAL}\x1b[0m")
    # Convert the dictionary items to a list and slice it
    items_list = list(cache_ids.items())[splice_range:]
    # Gets the latest section of ids, from checkpoint to the end
    cache_ids = dict(items_list)
    for row in cache_ids:
        foto_status = cache_ids[row]["status_value"]
        str_pelanggan = cache_ids[row]["str_pelanggan"]
        if foto_status == 'True':
            Log_write(f'No.{nomer} Foto terdeteksi [Skipping] . . . ID : {str_pelanggan}')
            nomer+=1
            continue
        Log_write(f'No.{nomer} Mencari foto . . .')
        search_pelanggan(str_pelanggan)
        data_foto = lihat_foto(str_pelanggan,nomer)
        if data_foto == False:
            Log_write("--> Cache img updated [No image] ..")
            download_photo(data_foto,row,"False")
            update_cache_ids(row,"False")
        else:
            download_photo(data_foto,row,"True")
            update_cache_ids(row,"True")
        current_time = datetime.now().time()
        # Extract hour, minute, and second components
        hour = current_time.hour
        minute = current_time.minute
        second = current_time.second
        # Log_write("--> Workbook updated!")
        Log_write(f'--> {hour}:{minute}:{second}')
        checkpoint(row,nomer,str_pelanggan)
        Log_write(f"\x1b[1;96m>> Total left : {ROW_AKHIR-ROW_AWAL-nomer}\x1b[0m")
        nomer+=1
    save_photo()
    logout_akun()
    driver.quit()
    delete_temp()
    Log_write('Webdriver flush\nExiting . . .')
    Log_write('\x1b[1;92mAll done ...')
    Log_write(show_vers())
    exit()

if __name__ == '__main__':
    # save_photo()
    # exit()
    driver = start_web_dv()
    Log_write(f"{show_vers()}\n")
    while True:
        try:
            main()
        except Exception as e:
            Log_write(f"Got error : {e}")
            Log_write(f"\x1b[1;35m--> Relogin [Refreshing]\x1b[0m","warning")
            logout_akun()
            last_logout_time = time()
            continue