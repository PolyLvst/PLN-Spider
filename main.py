version___ = 'PLN Spider v2.2'
import threading
from dotenv import load_dotenv
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.firefox.options import Options
from openpyxl.drawing.image import Image
from selenium import webdriver
from datetime import datetime
from openpyxl import load_workbook
from time import sleep,time
import os
import json
import base64
import logging

load_dotenv()
find_this = os.environ
URL = find_this['URL']
USER = find_this['USER_LOGIN']
PASSWORD = find_this['PASSWORD']
EXCEL_PATH = find_this['EXCEL_PATH']
ROW_AWAL = int(find_this['ROW_AWAL'])
ROW_AKHIR = int(find_this['ROW_AKHIR'])
COL_ID = find_this['COL_ID']
COL_PHOTO = find_this['COL_PHOTO']
COL_STAT = find_this['COL_STAT']
BANYAK_PERCOBAAN = int(find_this['BANYAK_PERCOBAAN']) # -- Berapa kali untuk mencoba mencari foto saat internet tidak stabil
# -- Setting Foto --
desired_width = int(find_this['desired_width'])
desired_height = int(find_this['desired_height'])

# Convert huruf menjadi angka untuk index pelanggan
# col_id_num= utils.column_index_from_string(COL_ID)-1
# col_photo_num = utils.column_index_from_string(COL_PHOTO)-1

# Sleep timer
sleep_for_filter = 3
sleep_for_search = 2
sleep_for_timeout_foto = 15
sleep_relog = 60 # 1 Menit
sleep_retry_foto = 2
sleep_tombol_close_foto = 5
last_logout_time = time()

now__ = datetime.now()
now_time = now__.strftime('%d-%m-%Y_%H-%M-%S')
log_file_path = './logs/PLN-Spider-'+now_time+'.log'
with open("./DataSnapshots/loglastrunpath.json","w") as f:
    json.dump({"log_path":log_file_path},f)

excel_file_path = EXCEL_PATH
base64_foto_tidak_tersedia = find_this['base_64_foto_tidak_tersedia']

def show_vers():
    created_by = find_this['creator']
    return f'\x1b[1;96m>> Created by : {created_by}\n>> Github : https://github.com/PolyLvst\n\x1b[1;93m@ {version___}\x1b[0m'
# ------------- Selenium web driver ------------ #
def start_web_dv(profile="default"):
    options = Options()
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0"
    options.set_preference("general.useragent.override", user_agent)
    options.set_preference("network.trr.mode", 2)
    options.set_preference("network.trr.uri", "https://mozilla.cloudflare-dns.com/dns-query")
    profile_path = os.path.join(os.getenv('APPDATA'), 'Mozilla', 'Firefox', 'Profiles')
    # List all directories in the Profiles folder
    profiles = [d for d in os.listdir(profile_path) if os.path.isdir(os.path.join(profile_path, d))]
    path_prof = None 
    for path_p in profiles:
        if profile in path_p:
            path_prof = os.path.join(profile_path, path_p)
            break
    if path_prof:
        Log_write(f"-- Using {profile} profile")
        options.add_argument("-profile")
        options.add_argument(path_prof)
    driver = webdriver.Firefox(options=options)
    return driver

def Log_write(text,stat='info'):
    """Available params ->>\ndebug,info,warning,error,critical"""
    text = str(text)
    log_filename = log_file_path
    logging.basicConfig(filename=log_filename, filemode='w', format='%(asctime)s - %(levelname)s - %(message)s', level=logging.DEBUG)
    # Set the logging level for the selenium logger to WARNING
    logging.getLogger('selenium').setLevel(logging.WARNING)
    # Set the logging level for the webdriver logger to WARNING
    logging.getLogger('webdriver').setLevel(logging.WARNING)
    # Set the logging level, to prevent unwanted message showing in log file
    logging.getLogger('urllib3.connectionpool').setLevel(logging.WARNING)
    print(text)
    text = text.replace('\n',' ')
    # Map the level string to a logging level constant
    level_map = {'debug': logging.DEBUG,
                 'info': logging.INFO,
                 'warning': logging.WARNING,
                 'error': logging.ERROR,
                 'critical': logging.CRITICAL}
    log_level = level_map.get(stat.lower(), logging.INFO)
    logging.log(log_level,text)

def input_login(user,passw,btn):
    user.send_keys(USER)
    passw.send_keys(PASSWORD)
    btn.click()

def logout_akun():
    current_time = time()
    elapsed_time = current_time - last_logout_time
    Log_write(f"Time since last logout : {elapsed_time}s")
    if elapsed_time <= 30:
        Log_write("Logout cooldown ... [30s]","warning")
        for i in range(30,0,-1):
            sleep(1)
            if i%2 == 0:
                print(f"\r[{i}] (-_-) zzZ ",end="")
            elif i%3 == 0:
                print(f"\r[{i}] (^-^) Zzz ",end="")
            else:
                print(f"\r[{i}] (-_-) zZz ",end="")
        print("\n")
    try:
        element = WebDriverWait(driver,45).until(EC.presence_of_element_located((By.XPATH,f"//div[@class='GCMY5A5CON'][contains(.,'{USER}')]")))
        div_tombol = driver.find_element(By.XPATH,f"//div[@class='GCMY5A5CON'][contains(.,'{USER}')]")
        div_tombol.click()
        element = WebDriverWait(driver,45).until(EC.presence_of_element_located((By.XPATH,"//a[@class='GCMY5A5CKVB'][contains(., 'Logout')]")))
        tombol_logout = driver.find_element(By.XPATH,"//a[@class='GCMY5A5CKVB'][contains(., 'Logout')]")
        tombol_logout.click()
        Log_write("Logged out ... ")
    except:
        Log_write("Logout button not found","error")
        exit(1)

def delete_temp():
    folder = "./TempImages"
    for file_img in os.listdir(folder):
        Log_write(f"deleting -> {file_img}")
        os.remove(f"{folder}/{file_img}")

def click_sidebar():
    try:
        # Folder MONITORING DAN LAPORAN
        element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'x-widget-8_f-14')))
        sidebar_tusbung_parent = driver.find_element(By.ID,'x-widget-8_f-14')
        sidebar_tusbung = sidebar_tusbung_parent.find_element(By.CSS_SELECTOR,'img.GCMY5A5CFOB')
        sidebar_tusbung.click()
        # Document Info Pelaksanaan TUL
        element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'x-widget-8_m-19')))
        informasi_TUL_parent = sidebar_tusbung_parent.find_element(By.ID,'x-widget-8_m-19')
        informasi_TUL = informasi_TUL_parent.find_element(By.CSS_SELECTOR,'img.GCMY5A5CEOB')
        informasi_TUL.click()
    except:
        Log_write("Something went wrong [Sidebar not detected]")
        # exit(1)
        raise Exception

def search_pelanggan(id_pelanggan):
    try:
        input_pelanggan = driver.find_element('id','x-widget-19-input')
        input_pelanggan.clear()
        input_pelanggan.send_keys(id_pelanggan)
        # tombol_cari = driver.find_element('xpath',"//div[@class='GCMY5A5CON'][contains(.,'Cari')]")
        tombol_cari = driver.find_element('xpath','/html/body/div[2]/div/div/div[1]/div[2]/div[1]/div/div[2]/div/div[2]/div[2]/div[1]/div/div/div[1]/div/div[1]/div/div[5]/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[1]/img')
        tombol_cari.click()
        # element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, "(//div[@class='GCMY5A5CNHC'][contains(.,'NO WO')])[1]")))
        element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div[1]/div[2]/div[1]/div/div[2]/div/div[2]/div[2]/div[1]/div/div/div[1]/div/div[4]/div[1]/div/table/tbody/tr/td[1]/div')))
        sleep(sleep_for_search)
        # filter_tahun = driver.find_element('xpath',"(//div[@class='GCMY5A5CNHC'][contains(.,'NO WO')])[1]")
        filter_tahun = driver.find_element('xpath','/html/body/div[2]/div/div/div[1]/div[2]/div[1]/div/div[2]/div/div[2]/div[2]/div[1]/div/div/div[1]/div/div[4]/div[1]/div/table/tbody/tr/td[1]/div')
        attribut_filter = filter_tahun.get_attribute("class")
    except:
        Log_write("Something wrong happens [search_pelanggan]","error")
        # exit(1)
        raise Exception
    if "GCMY5A5CEIC" in attribut_filter:
        pass
    else:
        sleep(sleep_for_filter)
        try:
            filter_tahun.click()
            filter_tahun.click()
        except:
            Log_write("Waiting filter tahun ..")
            sleep(10)
            filter_tahun.click()
            filter_tahun.click()

def table_filter(idx_bulan=1):
    try:
        actions = ActionChains(driver)
        # idx_bulan = 1 artinya bulan terbaru atau bulan saat ini
        # anda bisa menyesuaikan sesuai kebutuhan jika ingin ke bulan sebelumnya
        element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR,'table.GCMY5A5CMIC')))
        table_element = driver.find_element(By.CSS_SELECTOR,'table.GCMY5A5CMIC')
        bulan_ = table_element.find_elements(By.TAG_NAME,'tr')
        bulan_pilihan = bulan_[idx_bulan]
        driver.execute_script("arguments[0].scrollIntoView();", bulan_pilihan)
        actions.move_to_element(bulan_pilihan).click().perform()
        return len(bulan_)
    except:
        Log_write("Table filter error [Not found]","error")
        raise Exception

def lihat_foto(id_pelanggan,nomer):
    try:
        try:
            img_button = driver.find_element('xpath',"(//div[@class='GCMY5A5CON'][contains(.,'Lihat Foto')])[1]")
            img_button.click()
        except:
            sleep(sleep_for_timeout_foto)
            img_button = driver.find_element('xpath',"(//div[@class='GCMY5A5CON'][contains(.,'Lihat Foto')])[1]")
            img_button.click()
    except:
        Log_write("Popup detected, trying to close it","warning")
        try:
            popup = WebDriverWait(driver,40).until(EC.presence_of_element_located(By.XPATH,"//div[@class='GCMY5A5CON'][contains(.,'OK')]"))
            popup = driver.find_element('xpath',"//div[@class='GCMY5A5CON'][contains(.,'OK')]")
            popup.click()
            Log_write("Popup closed")
        except:
            Log_write("Failed to close it ... ")
            logout_akun()
            # exit(1)
            raise Exception
    # Wait for the image element to be visible
    trying = 1
    final_image_source = ''
    while True:
        Log_write(f'--> ID : {id_pelanggan}, Percobaan ke : {trying}')
        # will exit after +1 trying
        if trying > BANYAK_PERCOBAAN:
            # if trying >= BANYAK_PERCOBAAN+1:
            Log_write('--> Bad connection [Logout & Exit]','error')
            logout_akun()
            # exit(1)
            raise Exception
            # Log_write('--> Bad connection [Trying one last time again]','warning')
        try:
            image_pojok_kiri_bawah = WebDriverWait(driver,40).until(EC.presence_of_element_located((By.XPATH,'/html/body/div[5]/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div[2]/div[1]')))

        except TimeoutException as e:
            Log_write(f"--> Timeout exception occurred: [Waiting] trying again",'warning')
            # Log_write(f"--> e : {e}",'warning')
            sleep(sleep_for_timeout_foto)
            try:
                table_filter()
                Log_write(f'No.{nomer} Mencari foto [Retrying] . . .')
                data_fotoX = lihat_foto(id_pelanggan,nomer)
                # Log_write(f'String data foto : {data_fotoX}')
                return data_fotoX
            except Exception as ef:
                logout_akun()
                raise Exception
        # Pojok kiri bawah
        img_div_parent = driver.find_element('xpath','/html/body/div[5]/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div[2]/div[1]')
        img_element_1 = img_div_parent.find_element(By.CSS_SELECTOR,'img.gwt-Image')
        # Pojok kanan bawah
        img_div_parent = driver.find_element('xpath','/html/body/div[5]/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/div/div[2]/div[2]/div[2]/div[1]')
        img_element_2 = img_div_parent.find_element(By.CSS_SELECTOR,'img.gwt-Image')
        # Pojok kiri atas
        img_div_parent = driver.find_element('xpath','/html/body/div[5]/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/div/div[1]/div[1]/div[2]/div[1]')
        img_element_3 = img_div_parent.find_element(By.CSS_SELECTOR,'img.gwt-Image')
        # Pojok kanan atas
        img_div_parent = driver.find_element('xpath','/html/body/div[5]/div[2]/div[1]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/div/div[1]/div[2]/div[2]/div[1]')
        img_element_4 = img_div_parent.find_element(By.CSS_SELECTOR,'img.gwt-Image')

        wait_src = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "img")))
        image_source_1 = img_element_1.get_attribute("src")
        image_source_2 = img_element_2.get_attribute("src")
        image_source_3 = img_element_3.get_attribute("src")
        image_source_4 = img_element_4.get_attribute("src")
        if image_source_1 and image_source_2 and image_source_3 and image_source_4:
            if check_photo(image_source_1) == True:
                Log_write('--> Menggunakan foto pojok kiri bawah')
                final_image_source = image_source_1
                break
            elif check_photo(image_source_2) == True:
                Log_write('--> Menggunakan foto pojok kanan bawah')
                final_image_source = image_source_2
                break
            elif check_photo(image_source_3) == True:
                Log_write('--> Menggunakan foto pojok kiri atas')
                final_image_source = image_source_3
                break
            elif check_photo(image_source_4) == True:
                Log_write('--> Menggunakan foto pojok kanan atas')
                final_image_source = image_source_4
                break
            else:
                Log_write('--> Foto tidak tersedia','error')
                final_image_source = False
                break
        else:
            sleep(sleep_retry_foto)
        trying+=1
        
    tombol_close_parent = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[1]/div/div/div/div[2]")))
    tombol_close_parent = driver.find_element('xpath','/html/body/div[5]/div[1]/div/div/div/div[2]')
    tombol_close = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.GCMY5A5CCP.GCMY5A5CIK.GCMY5A5CHEC")))
    try:
        # wait_tombol = WebDriverWait(driver, 15).until(EC.invisibility_of_element_located((By.CLASS_NAME, 'GCMY5A5CJJ')))
        tombol_close = tombol_close_parent.find_element(By.CSS_SELECTOR,'div.GCMY5A5CCP.GCMY5A5CIK.GCMY5A5CHEC')
        tombol_close.click()
    except ElementClickInterceptedException as e:
        # Handle the exception when the click is not clickable
        Log_write(f"--> Click action failed: [Waiting and trying again]",'warning')
        sleep(sleep_tombol_close_foto)
        try:
            tombol_close = tombol_close_parent.find_element(By.CSS_SELECTOR,'div.GCMY5A5CCP.GCMY5A5CIK.GCMY5A5CHEC')
            tombol_close.click()
        except:
            Log_write("Failed to close img","error")
            logout_akun()
            # exit(1)
            raise Exception
    return final_image_source

def check_photo(source):
    data_url = source
    # Extract the base64-encoded image data from the Data URL
    image_data = data_url.split(",")[1]
    if image_data == base64_foto_tidak_tersedia:
        return False
    else:
        return True

def cache_photo(source,cur_pos,status_value):
    "cur pos = no-1"
    cache_img = "./DataSnapshots/cache_img.json"
    data_url = source
    # Extract the base64-encoded image data from the Data URL
    image_data = data_url.split(",")[1]
    image_format = data_url.split(";")[0].split(":")[1].split("/")[1]
    data = {}
    if os.path.exists(cache_img):
        with open(cache_img,"r") as f:
            data = json.load(f)
    with open(cache_img,"w") as f:
        data[cur_pos] = {"img":image_data,
                                 "format":image_format,
                                 "status_value":status_value}
        json.dump(data,f)
    Log_write("--> Cache img updated ..")

def save_photo():
    Log_write("Saving photos from cache ..")
    cache_img = "./DataSnapshots/cache_img.json"
    data = {}
    if os.path.exists(cache_img):
        with open(cache_img,"r") as f:
            data = json.load(f)
    else:
        Log_write("Something wrong [no cache img found]","error")
        exit(1)
    workbook = load_workbook(EXCEL_PATH)
    worksheet = workbook.active
    starts_row = ROW_AWAL
    for customer_id in data:
        Log_write(f"--> Saving {customer_id} ..")
        cur_pos = starts_row
        foto_cell=worksheet[f'{COL_PHOTO}{cur_pos}']
        image_data = data[customer_id]["img"]
        image_format = data[customer_id]["format"]
        status_value = data[customer_id]["status_value"]
        # Decode the base64 data into bytes
        image_bytes = base64.b64decode(image_data)
        
        # Save the image to a file
        image_path = f"./TempImages/tempimage-{customer_id}.{image_format}"
        with open(image_path, "wb") as file:
            file.write(image_bytes)

        img = Image(image_path)
        img.width = desired_width
        img.height = desired_height
        # add to worksheet and anchor next to cells
        worksheet.add_image(img, f'{COL_PHOTO}{cur_pos}')
        foto_cell.value = status_value
        starts_row += 1
    workbook.save(EXCEL_PATH)
    Log_write("--> Workbook updated!")
    workbook.close()

def search_past_image(past_month,id_pel,nomer):
    Log_write('--> Mencari foto di bulan sebelumnya')
    # Start from 2 because 1 is current month or the latest top of the table
    for i in range(2,past_month):
        Log_write(f'--> [{i-1}] Bulan sebelumnya')
        table_filter(i)
        d_foto = lihat_foto(id_pel,nomer)
        if d_foto == False:
            continue
        else:
            return d_foto
    Log_write(f'--> Pelanggan ini tidak mempunyai foto sama sekali','warning')
    return False

def check_folders():
    folder_doc = './Document'
    folder_temp_img = './TempImages'
    Log_write('Checking folder ... ')
    if os.path.exists(folder_doc) and os.path.exists(folder_temp_img):
        Log_write(f'Folder OK')
        return
    if not os.path.exists(folder_doc) and not os.path.exists(folder_temp_img):
        os.makedirs(folder_doc)
        os.makedirs(folder_temp_img)
        Log_write(f'Folder created : {folder_doc} & {folder_temp_img}','warning')
        return
    if not os.path.exists(folder_doc):
        os.makedirs(folder_doc)
        Log_write(f'Folder created : {folder_doc}','warning')
        return
    if not os.path.exists(folder_temp_img):
        os.makedirs(folder_temp_img)
        Log_write(f'Folder created : {folder_temp_img}','warning')
        return
    Log_write('Something went wrong ... [Folder related]','error')
    exit(1)

def check_status():
    workbook = load_workbook(EXCEL_PATH)
    worksheet = workbook.active
    status = worksheet[f'{COL_STAT}{ROW_AWAL}']
    if status.value == 'WORKING':
        Log_write('Working flag detected [Already been checked] ..')
        return
    end_row = ROW_AKHIR+1
    nomer = 0
    Log_write('Cleaning COL_PHOTO residue .. [Leftover flags]')
    for row in range(ROW_AWAL,end_row):
        nomer+=1
        foto_cell = worksheet[f'{COL_PHOTO}{row}']
        foto_cell.value = ''
    status.value = 'WORKING'
    Log_write('Workbook is clean and ready to be used [First time checkup]')
    workbook.save(EXCEL_PATH)
    Log_write("--> Workbook updated!")
    workbook.close()
    return

def remove_working_flag():
    workbook = load_workbook(excel_file_path)
    worksheet = workbook.active
    status = worksheet[f'{COL_STAT}{ROW_AWAL}']
    status.value = ''
    workbook.save(EXCEL_PATH)
    workbook.close()
    Log_write("--> Workbook updated! removed WORKING flag")

def clean_old_files(path_to):
    max_age_seconds = 3 * 24 * 60 * 60
    old_files = []
    for file_path in os.listdir(path_to):
        file_path = os.path.join(path_to,file_path)
        file_stat = os.stat(file_path)
        current_time = time()
        # Calculate the age of the file in seconds
        file_age_seconds = current_time - file_stat.st_mtime
        # Compare the age with the maximum allowed age
        if file_age_seconds > max_age_seconds:
            # File is older than 3 days
            old_files.append(file_path)
    if not old_files:
        return
    Log_write("Old files found in DataSnapshots ..")
    print("Use p for delete on prompt")
    cond = input("Delete all old files in DataSnapshots? [y/n/p] ").lower()
    if cond == "n":
        Log_write("No files deleted ..")
        return
    if cond != "y" and cond != "p":
        Log_write("Invalid choice")
        exit(0)
    for file in old_files:
        # prompt mode if cond not "y"
        if cond != "y":
            inp = input(f"Delete {file} ? [y/n] ").lower()
            if inp == "n":
                Log_write(f"Skipping {file} ..")
                continue
        # if cond == "y" then go delete all file without prompt
        os.remove(file)
        Log_write(f"{file} has been deleted as it's more than 3 days old.","warning")

def checkpoint(row,no,id_pel):
    checkpoint = './DataSnapshots/checkpoint.json'
    data = {}
    # with open(checkpoint,'r') as f:
    #     data = json.load(f)
    data.update({'row_awal':ROW_AWAL,'checkpoint':{'no':no,'row_checkpoint':row,'id':id_pel}})
    with open(checkpoint,'w') as f:
        json.dump(data,f)
        Log_write('Updated checkpoint ..')

def ask_checkpoint():
    data = {}
    nomer = 1
    row_awal = ROW_AWAL
    checkpoint = './DataSnapshots/checkpoint.json'
    if not os.path.exists('./DataSnapshots'):
        os.mkdir('./DataSnapshots')
    clean_old_files('./DataSnapshots')
    if os.path.exists(checkpoint):
        Log_write('--> Checkpoint found, using value from checkpoint')
        with open(checkpoint,'r') as f:
            try:
                data:dict = json.load(f)
            except:
                Log_write('No value detected, fallback default')
                return nomer,row_awal
        check_idx:dict = data['checkpoint']
        nomer = check_idx.get('no')
        row_awal = check_idx.get('row_checkpoint')
        return nomer,row_awal
    else:
        data = {'row_awal':ROW_AWAL,'checkpoint':{'no':1,'row_checkpoint':ROW_AWAL,'id':'first time run'}}
        with open(checkpoint,'w') as f:
            json.dump(data,f)
        Log_write('Init checkpoint ..')
        return nomer,row_awal

def get_cached_ids():
    ids_path = "./DataSnapshots/cached_ids.json"
    data = {}
    if os.path.exists(ids_path):
        Log_write("Using cached ids ..")
        with open(ids_path,"r") as f:
            data = json.load(f)
        return data
    workbook = load_workbook(excel_file_path)
    worksheet = workbook.active
    end_row = ROW_AKHIR+1
    for row in range(ROW_AWAL,end_row):
        id_pelanggan = worksheet[f'{COL_ID}{row}']
        str_pelanggan = id_pelanggan.value
        data[f"no-{str_pelanggan}"] = {"str_pelanggan":str_pelanggan,
                                       "status_value":"Init_state"}
    with open(ids_path,"w") as f:
        json.dump(data,f)
    Log_write("Init cached ids ..")
    workbook.close()
    return data

def update_cache_ids(no_id,status_given):
    "no_id = no-1"
    ids_path = "./DataSnapshots/cached_ids.json"
    data = {}
    if os.path.exists(ids_path):
        with open(ids_path,"r") as f:
            data = json.load(f)
        data[no_id]["status_value"] = status_given
        with open(ids_path,"w") as f:
            json.dump(data,f)
        Log_write("Status value updated ..")
    else:
        Log_write("Something wrong happens [cached_ids not found]","error")
        exit(1)

def listen_for_input(stop_event):
    while True:
        user_input = input("Enter 'q' to quit: \n")
        if user_input.lower() == 'q':
            print("Quitting the program.")
            stop_event.set()  # Set the event to signal the loop job to stop
            break

# ------------------- MAIN PROGRAM ------------------
# Dapat juga berfungsi sebagai module
def main(stop_event):
    check_folders()
    check_status()
    driver.get(URL)
    element = WebDriverWait(driver, 35).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.GCMY5A5CFN')))
    input_login_user = driver.find_element('id','x-widget-1-input')
    input_login_password = driver.find_element('id','x-widget-2-input')
    button_login = driver.find_element('xpath','/html/body/div[3]/div[2]/div[1]/div/div/div[2]/div/div/div/div/table/tbody/tr[2]/td[2]/div/div/table/tbody/tr/td[1]/img')
    if input_login_user and input_login_password:
        input_login(input_login_user,input_login_password,button_login)
    else:
        Log_write('Something went wrong ! [User input not found]','error')
        driver.quit()
        exit(1)
    click_sidebar()
    Log_write('Logged in')
    nomer,row_awal = ask_checkpoint()
    cache_ids = get_cached_ids()
    if nomer == 1:
        splice_range = 0
    else:
        splice_range = nomer-1
    # Convert the dictionary items to a list and slice it
    items_list = list(cache_ids.items())[splice_range:]
    # Gets the latest section of ids, from checkpoint to the end
    cache_ids = dict(items_list)
    for row in cache_ids:
        if stop_event.is_set():
            break
        foto_status = cache_ids[row]["status_value"]
        str_pelanggan = cache_ids[row]["str_pelanggan"]
        if foto_status == 'True':
            Log_write(f'No.{nomer} Foto terdeteksi [Skipping] . . . ID : {str_pelanggan} ROW : {row_awal}')
            nomer+=1
            continue
        if foto_status == 'Past':
            Log_write(f'No.{nomer} Foto terdeteksi, foto bulan lalu [Skipping] . . . ID : {str_pelanggan} ROW : {row_awal}')
            nomer+=1
            continue
        
        Log_write(f'No.{nomer} Mencari foto . . .')
        search_pelanggan(str_pelanggan)
        banyak_bulan = table_filter()
        data_foto = lihat_foto(str_pelanggan,nomer)
        # Jika tidak menemukan foto pada bulan ini maka memakai bulan sebelumnya
        if data_foto == False:
            data_foto = search_past_image(banyak_bulan,str_pelanggan,nomer)
            # Jika tetap tidak menemukan sama sekali foto pada bulan sebelumnya maka gunakan foto tidak tersedia
            if data_foto == False:
                # foto_cell.value = 'False'
                cache_photo(f'data:image/jpg;base64,{base64_foto_tidak_tersedia}',row,"False")
                update_cache_ids(row,"False")
            else:
                # foto_cell.value = 'Past'
                cache_photo(data_foto,row,"Past")
                update_cache_ids(row,"Past")
        else:
            # foto_cell.value = save_photo(data_foto,row)
            cache_photo(data_foto,row,"True")
            update_cache_ids(row,"True")
        # try:
        current_time = datetime.now().time()
        # Extract hour, minute, and second components
        hour = current_time.hour
        minute = current_time.minute
        second = current_time.second
        # Log_write("--> Workbook updated!")
        Log_write(f'--> {hour}:{minute}:{second}')
        checkpoint(row,nomer,str_pelanggan)
        nomer+=1
        # except Exception as e:
        #     Log_write(f"An error occurred while saving the workbook: {e}",'error')
    save_photo()
    remove_working_flag()
    logout_akun()
    driver.quit()
    delete_temp()
    Log_write('Webdriver flush\nExiting . . .')
    Log_write('\x1b[1;92mAll done ...')
    Log_write(show_vers())
    exit()

if __name__ == '__main__':
    driver = start_web_dv()
    Log_write(f"{show_vers()}\n")
    while True:
        try:
            # Create a threading.Event object
            stop_event = threading.Event()
            
            # Create threads for input listener and main job
            input_thread = threading.Thread(target=listen_for_input, args=(stop_event,))
            main_thread = threading.Thread(target=main, args=(stop_event,))
            
            # Start the threads
            input_thread.start()
            main_thread.start()
            
            # Wait for the threads to finish
            input_thread.join()
            main_thread.join()

        except Exception:
            Log_write(f"\x1b[1;35m--> Relogin [Refreshing]\x1b[0m","warning")
            logout_akun()
            last_logout_time = time()
            continue